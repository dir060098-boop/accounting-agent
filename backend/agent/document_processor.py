import fitz  # PyMuPDF
import openpyxl
import io
from fastapi import UploadFile


async def extract_text_from_pdf(file: UploadFile) -> str:
    content = await file.read()
    doc = fitz.open(stream=content, filetype="pdf")
    text = ""
    for page in doc:
        text += page.get_text()
    doc.close()
    await file.seek(0)
    return text.strip()


async def extract_text_from_excel(file: UploadFile) -> str:
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    text_parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        text_parts.append(f"Лист: {sheet_name}")
        for row in ws.iter_rows(values_only=True):
            row_text = " | ".join(str(cell) for cell in row if cell is not None)
            if row_text.strip():
                text_parts.append(row_text)
    await file.seek(0)
    return "\n".join(text_parts)


async def extract_text(file: UploadFile) -> str:
    filename = file.filename.lower()
    if filename.endswith(".pdf"):
        return await extract_text_from_pdf(file)
    elif filename.endswith((".xlsx", ".xls")):
        return await extract_text_from_excel(file)
    elif filename.endswith(".txt"):
        content = await file.read()
        await file.seek(0)
        return content.decode("utf-8", errors="ignore")
    else:
        raise ValueError(f"Неподдерживаемый формат файла: {file.filename}")
